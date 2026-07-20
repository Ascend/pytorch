"""
从 Excel 中提取 Status 为 Done 的文件，匹配流水线产出的 shard_*_cases.json，
生成新的 Excel，前3列保持一致，第4列为 nodeid，第5列为执行结果。

用法:
    python generate_report.py \
        --excel "Test Class Refactoring Tracker.xlsx" \
        --reports-dir all-test-reports \
        --output output/done_cases_report.xlsx \
        [--v271-failed V2.7.1_all_failed_testcases.xlsx] \
        [--ops ops.txt]
"""
import argparse
import os
import json
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# 需要处理的 Sheet（排除 README 类）
SKIP_SHEETS = {"README", "README(Bak)"}

# ========== 日志过滤：不支持模式 ==========
EXCEPTION_INPUT_DTYPES = ["float64", "complex"]

UNSUPPORTED_PATTERNS = [
    # 硬件不支持
    ("not implemented for DT_COMPLEX",),
    ("not implemented for DT_DOUBLE",),
    *[("Exception: Caused by " + t + " input at index", "dtype=torch." + d)
      for t in ("sample", "error", "reference") for d in EXCEPTION_INPUT_DTYPES],
    ("cannot be larger than 8 dimensions",),
    ("Jiterator is only supported on CUDA",),
    ("RuntimeError: Only contiguous_format or preserve_format is supported.",),
    ("RuntimeError: NPU contiguous operator only supported contiguous memory format.",),
    # 算子支持
    ("aten::_scaled_dot_product_flash_attention",),
    ("aten::_fill_mem_eff_dropout_mask_",),
]


def load_ops(ops_path):
    """加载算子列表"""
    ops = set()
    if not ops_path or not os.path.exists(ops_path):
        if ops_path:
            print(f"  Warning: {ops_path} not found, op matching disabled")
        return ops
    with open(ops_path, "r", encoding="utf-8") as f:
        for line in f:
            name = line.strip()
            if name:
                ops.add(name)
    return ops


def check_unsupported_pattern(message):
    """检查报错日志是否匹配不支持模式"""
    if not message:
        return None
    for pattern_tuple in UNSUPPORTED_PATTERNS:
        if all(p in message for p in pattern_tuple):
            return " + ".join(pattern_tuple)
    return None


def check_ops_in_message(message, ops):
    """检查报错日志是否包含特定算子"""
    if not message or not ops:
        return None
    for op in ops:
        if op in message:
            return op
    return None


def classify_case(status, message, ops):
    """
    对用例进行分类，返回 (是否不支持, 不支持原因)
    """
    if status in ("passed", "skipped"):
        return "否", ""

    pattern = check_unsupported_pattern(message)
    if pattern:
        return "是", pattern
    op = check_ops_in_message(message, ops)
    if op:
        return "是", op
    return "否", ""


def load_all_cases(reports_dir):
    """从流水线产出的 shard_*_cases.json 中加载用例数据，按 file 路径索引"""
    cases_by_file = defaultdict(list)

    if not os.path.isdir(reports_dir):
        print(f"  Warning: reports dir not found: {reports_dir}")
        return dict(cases_by_file)

    for fname in sorted(os.listdir(reports_dir)):
        if not fname.endswith("_cases.json"):
            continue
        filepath = os.path.join(reports_dir, fname)
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            print(f"  Warning: failed to read {fname}: {e}")
            continue

        for case in data.get("cases", []):
            file_key = case.get("file", "")
            if file_key:
                cases_by_file[file_key].append(
                    (case["nodeid"], case["status"], case.get("message", ""))
                )

    return dict(cases_by_file)


def normalize_path(file_path):
    """统一文件路径格式"""
    if file_path is None:
        return None
    path = str(file_path).strip()
    if path.startswith("test/"):
        return path
    return "test/" + path


def match_file(excel_path, cases_by_file):
    """尝试匹配 Excel 中的文件路径到 cases 中的文件路径"""
    if excel_path is None:
        return None

    norm = normalize_path(excel_path)

    if norm in cases_by_file:
        return norm
    if norm.startswith("test/") and norm[5:] in cases_by_file:
        return norm[5:]
    if not norm.startswith("test/") and ("test/" + norm) in cases_by_file:
        return "test/" + norm

    for case_path in cases_by_file:
        if case_path.endswith(norm) or norm.endswith(case_path):
            return case_path

    return None


def load_v271_failed(v271_path):
    """从 V2.7.1_all_failed_testcases.xlsx 加载失败用例，按 nodeid 索引"""
    v271_map = {}
    if not v271_path or not os.path.exists(v271_path):
        if v271_path:
            print(f"  Warning: {v271_path} not found, V2.7.1 comparison disabled")
        return v271_map

    wb = openpyxl.load_workbook(v271_path)
    ws = wb["失败用例明细"]
    for row_idx in range(2, ws.max_row + 1):
        nodeid = ws.cell(row=row_idx, column=4).value
        status = ws.cell(row=row_idx, column=5).value
        if nodeid:
            v271_map[str(nodeid).strip()] = str(status).strip() if status else ""
    wb.close()
    return v271_map


def main():
    parser = argparse.ArgumentParser(
        description="从 Excel + 流水线用例结果生成分析报告"
    )
    parser.add_argument("--excel", required=True, help="Test Class Refactoring Tracker.xlsx 路径")
    parser.add_argument("--reports-dir", required=True, help="流水线产出的 shard_*_cases.json 所在目录")
    parser.add_argument("--output", required=True, help="输出 Excel 路径")
    parser.add_argument("--v271-failed", default=None, help="V2.7.1 失败用例 Excel（可选）")
    parser.add_argument("--ops", default=None, help="算子列表文件（可选）")
    args = parser.parse_args()

    excel_path = args.excel
    reports_dir = args.reports_dir
    output_path = args.output
    v271_path = args.v271_failed
    ops_path = args.ops

    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    print("加载用例数据...")
    cases_by_file = load_all_cases(reports_dir)
    print(f"  共加载 {len(cases_by_file)} 个文件的用例数据")

    print("加载 V2.7.1 失败用例...")
    v271_map = load_v271_failed(v271_path)
    print(f"  共加载 {len(v271_map)} 条 V2.7.1 失败用例")

    print("加载算子列表...")
    ops = load_ops(ops_path)
    print(f"  共加载 {len(ops)} 个算子")

    wb = openpyxl.load_workbook(excel_path)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    status_colors = {
        "passed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "skipped": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "error": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "timeout": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    total_matched_files = 0
    total_matched_cases = 0
    total_not_found = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        out_ws = out_wb.create_sheet(title=sheet_name)

        headers = []
        for col_idx in range(1, 4):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(val if val else "")
        headers.extend(["nodeid", "执行结果", "2.7.1 失败", "报错日志", "不支持", "不支持原因"])

        for col_idx, header in enumerate(headers, 1):
            cell = out_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        out_row = 2
        sheet_matched_files = 0
        sheet_matched_cases = 0
        sheet_not_found = 0

        last_c1 = None
        last_c2 = None
        group_merge_start = None
        group_c1 = None
        group_c2 = None

        def flush_group_merge():
            nonlocal group_merge_start, group_c1, group_c2
            if group_merge_start is not None:
                merge_end = out_row - 1
                if merge_end > group_merge_start:
                    out_ws.merge_cells(start_row=group_merge_start, start_column=1, end_row=merge_end, end_column=1)
                    out_ws.merge_cells(start_row=group_merge_start, start_column=2, end_row=merge_end, end_column=2)
                    out_ws.cell(row=group_merge_start, column=1).alignment = Alignment(vertical="center")
                    out_ws.cell(row=group_merge_start, column=2).alignment = Alignment(vertical="center")
                group_merge_start = None
                group_c1 = None
                group_c2 = None

        for row_idx in range(2, ws.max_row + 1):
            c1_val = ws.cell(row=row_idx, column=1).value
            c2_val = ws.cell(row=row_idx, column=2).value
            if c1_val is not None:
                last_c1 = c1_val
            if c2_val is not None:
                last_c2 = c2_val

            status_cell = ws.cell(row=row_idx, column=4).value
            if status_cell is None:
                continue

            status_str = str(status_cell).strip()
            if "Done" not in status_str:
                continue

            file_path = ws.cell(row=row_idx, column=3).value
            if file_path is None:
                continue

            if group_c1 != last_c1 or group_c2 != last_c2:
                flush_group_merge()
                group_merge_start = out_row
                group_c1 = last_c1
                group_c2 = last_c2

            matched_path = match_file(file_path, cases_by_file)

            if matched_path is None:
                for col_idx, val in [(1, last_c1), (2, last_c2), (3, file_path)]:
                    cell = out_ws.cell(row=out_row, column=col_idx, value=val if val else "")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                out_ws.cell(row=out_row, column=4, value="(未匹配)")
                out_ws.cell(row=out_row, column=5, value="N/A")
                for col_idx in range(6, 10):
                    out_ws.cell(row=out_row, column=col_idx, value="")
                for col_idx in range(1, 10):
                    out_ws.cell(row=out_row, column=col_idx).border = thin_border
                out_row += 1
                sheet_not_found += 1
                continue

            sheet_matched_files += 1
            cases = cases_by_file[matched_path]

            for nodeid, case_status, message in cases:
                for col_idx, val in [(1, last_c1), (2, last_c2), (3, file_path)]:
                    cell = out_ws.cell(row=out_row, column=col_idx, value=val if val else "")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                cell = out_ws.cell(row=out_row, column=4, value=nodeid)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                cell = out_ws.cell(row=out_row, column=5, value=case_status)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if case_status in status_colors:
                    cell.fill = status_colors[case_status]

                v271_status = v271_map.get(nodeid, "")
                cell = out_ws.cell(row=out_row, column=6, value="是" if v271_status else "")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if v271_status:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                cell = out_ws.cell(row=out_row, column=7, value=message if message else "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                is_unsupported, unsupported_reason = classify_case(case_status, message, ops)
                cell = out_ws.cell(row=out_row, column=8, value=is_unsupported)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_unsupported == "是":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                cell = out_ws.cell(row=out_row, column=9, value=unsupported_reason)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                out_row += 1
                sheet_matched_cases += 1

        flush_group_merge()

        out_ws.column_dimensions["A"].width = 18
        out_ws.column_dimensions["B"].width = 22
        out_ws.column_dimensions["C"].width = 55
        out_ws.column_dimensions["D"].width = 90
        out_ws.column_dimensions["E"].width = 14
        out_ws.column_dimensions["F"].width = 14
        out_ws.column_dimensions["G"].width = 80
        out_ws.column_dimensions["H"].width = 16
        out_ws.column_dimensions["I"].width = 50

        for r in range(1, out_row):
            out_ws.row_dimensions[r].height = 20

        print(f"  [{sheet_name}] 匹配文件: {sheet_matched_files}, 用例数: {sheet_matched_cases}, 未匹配: {sheet_not_found}")
        total_matched_files += sheet_matched_files
        total_matched_cases += sheet_matched_cases
        total_not_found += sheet_not_found

    out_wb.save(output_path)
    print(f"\n完成！输出文件: {output_path}")
    print(f"总计: 匹配文件 {total_matched_files} 个, 用例 {total_matched_cases} 条, 未匹配 {total_not_found} 个")


if __name__ == "__main__":
    main()
