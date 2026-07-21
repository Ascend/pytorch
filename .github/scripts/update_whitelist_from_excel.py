#!/usr/bin/env python3
"""
从 Excel 中提取测试文件路径，替换 case_paths_ci.yml 的白名单，清空黑名单。

Excel 结构:
  - 多个 sheet: Core, Tensor, Distributed, Graph, Math, Quantization, Utils
  - 第 3 列 (File) 包含测试文件路径，如 test/nn/attention/test_fa3.py
  - 第 1 行是表头，数据从第 2 行开始

用法:
    python update_whitelist_from_excel.py <excel_path> <case_paths_yml_path>
"""

import argparse
import sys
from pathlib import Path
from typing import List

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("Error: pyyaml is required. Install with: pip install pyyaml")
    sys.exit(1)


# 需要读取的 sheet 名称
SHEET_NAMES = ["Core", "Tensor", "Distributed", "Graph", "Math", "Quantization", "Utils"]

# File 列索引 (1-based)
FILE_COLUMN = 3
# Status 列索引 (1-based)
STATUS_COLUMN = 4


def extract_test_files(excel_path: Path, status_filter: str = None) -> List[str]:
    """从 Excel 各 sheet 中提取测试文件路径（去重排序）。
    
    Args:
        excel_path: Excel 文件路径
        status_filter: 只提取匹配此 Status 的文件，None 表示不过滤
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    test_files = set()

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        count = 0
        skipped = 0
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row, FILE_COLUMN).value
            if not cell_value or not isinstance(cell_value, str):
                continue
            path = cell_value.strip()
            if not path:
                continue

            if status_filter:
                status = ws.cell(row, STATUS_COLUMN).value
                status_str = status.strip() if isinstance(status, str) else ""
                if status_filter not in status_str:
                    skipped += 1
                    continue

            test_files.add(path)
            count += 1

        if status_filter:
            print(f"  {sheet_name}: {count} files extracted (skipped {skipped} non-matching)")
        else:
            print(f"  {sheet_name}: {count} files extracted")

    wb.close()

    # 排序：先按目录层级，再按文件名
    sorted_files = sorted(test_files)
    return sorted_files


def update_case_paths_yml(yml_path: Path, whitelist: List[str]) -> None:
    """替换 whitelist，清空 blacklist。"""
    with open(yml_path, "r", encoding="utf-8") as f:
        content = yaml.safe_load(f)

    if content is None:
        content = {}

    content["whitelist"] = whitelist
    content["blacklist"] = []

    with open(yml_path, "w", encoding="utf-8") as f:
        yaml.dump(content, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

    print(f"\nUpdated {yml_path}:")
    print(f"  whitelist: {len(whitelist)} entries")
    print(f"  blacklist: 0 entries (cleared)")


def main():
    parser = argparse.ArgumentParser(
        description="从 Excel 提取测试文件路径，替换 case_paths_ci.yml 白名单"
    )
    parser.add_argument(
        "excel_path",
        type=str,
        help="Excel 文件路径",
    )
    parser.add_argument(
        "yml_path",
        type=str,
        help="case_paths_ci.yml 文件路径",
    )
    parser.add_argument(
        "--status",
        type=str,
        default=None,
        help="只提取匹配此 Status 的文件（如 'Done'），不指定则提取全部",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_path).resolve()
    yml_path = Path(args.yml_path).resolve()

    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)
    if not yml_path.exists():
        print(f"Error: YAML file not found: {yml_path}")
        sys.exit(1)

    print(f"Reading: {excel_path}")
    if args.status:
        print(f"Status filter: '{args.status}'")
    test_files = extract_test_files(excel_path, status_filter=args.status)
    print(f"\nTotal unique test files: {len(test_files)}")

    update_case_paths_yml(yml_path, test_files)

    # 打印前 10 条预览
    print("\nFirst 10 entries:")
    for f in test_files[:10]:
        print(f"  - {f}")
    if len(test_files) > 10:
        print(f"  ... and {len(test_files) - 10} more")


if __name__ == "__main__":
    main()
